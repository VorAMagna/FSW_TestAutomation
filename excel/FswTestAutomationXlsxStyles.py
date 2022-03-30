from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, \
    PatternFill


class FswTestAutomationXlsxStyles:
    """Singleton class type, needs to be initialized with the get_instance
    methdod, because the excel styles can only be constructed once!"""
    __instance__ = None

    def __init__(self):
        if FswTestAutomationXlsxStyles.__instance__ is None:
            FswTestAutomationXlsxStyles.__instance__ = self
            # Styles get initialized here
            self.default_style = NamedStyle(name="Default")

            self.test_case_failed_style = NamedStyle(name="failed_style")
            self.test_case_failed_style.font = Font(bold=True)
            self.test_case_failed_style.border = Border(
                bottom=Side(border_style="thin"))
            self.test_case_failed_style.alignment = Alignment(
                horizontal="center",
                vertical="center")
            self.test_case_failed_style.fill = PatternFill(
                fgColor="BE0000",
                fill_type="solid")

            self.test_case_successful_style = NamedStyle(
                name="successful_style")
            self.test_case_successful_style.fill = PatternFill(
                fgColor="2EB82E",
                fill_type="solid")
            self.test_case_successful_style.alignment = Alignment(
                horizontal="center",
                vertical="center")

            self.test_case_info_style = NamedStyle(name="info_style")
            self.test_case_info_style.alignment = Alignment(
                horizontal="center",
                vertical="center")

            self.top_style = NamedStyle(name="top_style")
            self.top_style.font = Font(bold=True)
            self.top_style.border = Border(bottom=Side(border_style="thin"))
            self.top_style.alignment = Alignment(horizontal="center",
                                                 vertical="center")
            self.top_style.fill = PatternFill(fgColor="00CCFF",
                                              fill_type="solid")

            self.over_limit_style = NamedStyle(name="over_limit_style")
            self.over_limit_style.fill = PatternFill(fgColor="FF3300",
                                                     fill_type="solid")
            self.over_limit_style.font = Font(bold=True)

            self.under_limit_style = NamedStyle(name="under_limit_style")
            self.under_limit_style.fill = PatternFill(fgColor="2EB82E",
                                                      fill_type="solid")
        else:
            raise Exception("You cannot create another SingletonGovt class")

    @staticmethod
    def get_instance():
        """
         Static method to fetch the current instance.
        """
        if not FswTestAutomationXlsxStyles.__instance__:
            FswTestAutomationXlsxStyles()
        return FswTestAutomationXlsxStyles.__instance__