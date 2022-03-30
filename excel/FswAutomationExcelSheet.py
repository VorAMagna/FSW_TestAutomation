from openpyxl.styles import Alignment
from global_config import WORD_WRAP_THRESHOLD


class FswAutomationExcelSheet:
    """Excel methods used for further excel reading/writing"""

    def __init__(self, access_type: str, workbook_ref, sheet_name):
        """
        Sets up a single excel sheet for reading or writing
        :param access_type: can be "read" or "write"
        :param workbook_ref: a reference of the workbook containing all the
        sheets
        :param sheet_name: the sheetname
        """
        self.column_index = 1
        self.row_index = 1
        if access_type == "write":
            self.sheet = workbook_ref.create_sheet(sheet_name)
        elif access_type == "read":
            self.sheet = workbook_ref[sheet_name]

    def auto_adjust_column_dimensions(self):
        """Adjusts the column/row dimentions so all the text fits"""
        for column_rows in self.sheet.iter_cols():
            max_width = 5
            max_height = 1
            for column_row in column_rows:

                if not isinstance(column_row.value, str):
                    continue

                column_row_adjusted_length = len(column_row.value)
                # Check if its a hyperlink, then crop to half size
                if "HYPERLINK" in column_row.value:
                    column_row_adjusted_length /= 2

                if column_row.value.count("\n") == 0 and \
                        column_row_adjusted_length > max_width:
                    max_width = column_row_adjusted_length

                # if the cell has \n in it, this will adjust the height of the
                # cell and activate wrap text
                if len(column_row.value) > WORD_WRAP_THRESHOLD:
                    column_row.alignment = Alignment(wrap_text=True)

                if column_row.value.count("\n") > (max_height + 1):
                    max_height = column_row.value.count("\n")

                    self.sheet.row_dimensions[
                        column_rows.index(column_row) + 1].height = \
                        max_height * 15

                    column_row.alignment = Alignment(wrap_text=True)

            self.sheet.column_dimensions[column_rows[0].column_letter].width \
                = int(max_width + 3)

    def format_cell_types(self):
        """Formats all the cell types accordingly (should be done at the end)"""
        self.sheet.auto_filter.ref = self.sheet.dimensions

    def new_row(self):
        """Moves the cell index at the start of the next row"""
        self.column_index = 1
        self.row_index += 1

    def new_column(self):
        """Moves the column to the right"""
        self.column_index += 1

    def write_to_cell(self, value):
        """Writes to the cell at the current position"""
        self.sheet.cell(row=self.row_index,
                        column=self.column_index).value = value

    def style_cell(self, style):
        """applies style to cell at current position"""
        self.sheet.cell(row=self.row_index,
                        column=self.column_index).style = style

    def get_cell_value(self, **kwargs):
        """Gets the value at the current cell
        can also get a specific cells value, if row_index and column_index
        is defined in kwargs"""

        if "row_index" in kwargs and kwargs['row_index'] != 0:
            row_index = kwargs['row_index']
        else:
            row_index = self.row_index
        if "column_index" in kwargs and kwargs['column_index'] != 0:
            column_index = kwargs['column_index']
        else:
            column_index = self.column_index

        ret_val = self.sheet.cell(row=row_index,
                                  column=column_index).value

        if "new_row" in kwargs and kwargs['new_row']:
            self.new_row()
        if "new_column" in kwargs and kwargs['new_column']:
            self.new_column()

        return ret_val

    def get_cell_style(self):
        """Get the style at the current cell"""
        return self.sheet.cell(row=self.row_index,
                               column=self.column_index).style

    def fill_cell(self, **kwargs):
        """Fills the cells at the current row/column,
        can also increment the row/column"""
        if "style" in kwargs:
            self.style_cell(kwargs['style'])
        if "value" in kwargs:
            self.write_to_cell((kwargs['value']))
        if "new_row" in kwargs and kwargs['new_row']:
            self.new_row()
        if "new_column" in kwargs and kwargs['new_column']:
            self.new_column()

    def read_cell(self, *args, **kwargs) -> dict:
        """Reads a cell and optionaly increments the row and column index"""
        ret_values = {}
        if "style" in args:
            ret_values['style'] = self.get_cell_style()
        if "value" in args:
            ret_values['value'] = self.get_cell_value()
        if "new_row" in kwargs and kwargs['new_row']:
            self.new_row()
        if "new_column" in kwargs and kwargs['new_column']:
            self.new_column()

    @staticmethod
    def format_hyperlink(url: str, text: str) -> str:
        """Formats text into a xlsx hyperlink"""
        return '=HYPERLINK("{}", "{}")' \
            .format(url, text)
