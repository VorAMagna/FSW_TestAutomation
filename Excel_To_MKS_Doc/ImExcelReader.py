from openpyxl import load_workbook

from Excel_To_MKS_Doc.config.field_name_maps import mapping as MAPS
from excel.FswAutomationExcelSheet import FswAutomationExcelSheet


class ImExcelReader(FswAutomationExcelSheet):
    """Reads the excel files"""

    def __init__(self, excel_file_path, sheet_name):
        self.header_row_index = 1
        self.workbook = load_workbook(filename=excel_file_path)
        super().__init__("read", self.workbook, sheet_name)
        # a list containing dictionaries with all the export information needed
        self.rm_export_list = list()
        self.fill_rm_export_list()

    def get_current_header_value(self) -> str:
        """Gets the header name above the current cell"""
        return self.get_cell_value(row_index=self.header_row_index)

    @staticmethod
    def get_according_mks_name(excel_name) -> str:
        """Gets the respective mks name from the MAPS list"""
        for map in MAPS:
            if map[0] == excel_name:
                return map[1]

    def fill_rm_export_list(self):
        """Extracts all the relevant data from the excel file into an export
        list with dictionaries"""

        # first relevant row
        self.row_index = EXCEL_FIRST_RELEVANT_ROW
        # needs to be initialized here because the max columns increase
        # as we read the last column
        max_columns = self.sheet.max_column
        max_rows = self.sheet.max_row

        # loops every column and row after the current row/column index
        while self.row_index <= max_rows:
            curr_item = {}
            while self.column_index <= max_columns:
                curr_item[
                    self.get_according_mks_name(
                        self.get_current_header_value())] = \
                    self.get_cell_value()
                self.new_column()
            self.new_row()
            self.rm_export_list.append(curr_item)

    def fill_export_list(self):
        """Fills the export list with the values in the excel file"""
        pass


EXCEL_FIRST_RELEVANT_ROW = 2
EXCEL_HEADER_ROW = 1